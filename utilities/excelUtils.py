import openpyxl


def create_excel_object(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return workbook, sheet


def get_row_count(file_path, sheet_name):
    workbook, sheet = create_excel_object(file_path=file_path, sheet_name=sheet_name)
    return sheet.max_row

def get_col_count(file_path, sheet_name):
    workbook, sheet = create_excel_object(file_path=file_path, sheet_name=sheet_name)
    return sheet.max_row


def read_data(file_path, sheet_name,row_num,col_num):
    workbook, sheet = create_excel_object(file_path=file_path, sheet_name=sheet_name)
    return sheet.cell(row_num,col_num).value


def write_data(file_path, sheet_name,row_num,col_num,data):
    workbook, sheet = create_excel_object(file_path=file_path, sheet_name=sheet_name)
    sheet.cell(row_num,col_num).value = data
    workbook.save(file_path)